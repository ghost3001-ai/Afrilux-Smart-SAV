import csv
from datetime import datetime, time, timedelta
from io import BytesIO, StringIO

from django.db.models import Avg, Count, Q
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

try:
    from openpyxl import Workbook
except ModuleNotFoundError:  # pragma: no cover - depends on local environment
    Workbook = None

from .models import Intervention, Product, Ticket, TicketFeedback
from .services import (
    OPEN_TICKET_STATUSES,
    compute_agent_performance_rows,
    compute_average_first_response_hours,
    compute_average_resolution_hours,
    scope_intervention_queryset,
    scope_ticket_queryset,
)


REPORT_DAILY = "journalier"
REPORT_WEEKLY = "hebdomadaire"
REPORT_MONTHLY = "mensuel"

REPORT_CHOICES = {
    REPORT_DAILY: "Rapport journalier",
    REPORT_WEEKLY: "Rapport hebdomadaire",
    REPORT_MONTHLY: "Rapport mensuel",
}


def _aware_bounds(start_date, end_date):
    start_dt = timezone.make_aware(datetime.combine(start_date, time.min))
    end_dt = timezone.make_aware(datetime.combine(end_date, time.min))
    return start_dt, end_dt


def _period_for(report_type, anchor_date=None):
    anchor_date = anchor_date or timezone.localdate()
    if report_type == REPORT_DAILY:
        start_date = anchor_date
        end_date = anchor_date + timedelta(days=1)
        label = anchor_date.strftime("%d/%m/%Y")
    elif report_type == REPORT_WEEKLY:
        start_date = anchor_date - timedelta(days=anchor_date.weekday())
        end_date = start_date + timedelta(days=7)
        label = f"Semaine du {start_date.strftime('%d/%m/%Y')}"
    elif report_type == REPORT_MONTHLY:
        start_date = anchor_date.replace(day=1)
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year + 1, month=1)
        else:
            end_date = start_date.replace(month=start_date.month + 1)
        label = start_date.strftime("%m/%Y")
    else:
        raise ValueError("Type de rapport inconnu.")
    start_dt, end_dt = _aware_bounds(start_date, end_date)
    return start_date, end_date, start_dt, end_dt, label


def _serialize_ticket(ticket):
    return {
        "reference": ticket.reference,
        "titre": ticket.title,
        "client": str(ticket.client),
        "equipement": ticket.product.serial_number if ticket.product else "",
        "statut": ticket.get_status_display(),
        "priorite": ticket.get_priority_display(),
        "deadline_sla": ticket.sla_deadline.isoformat() if ticket.sla_deadline else "",
        "technicien": str(ticket.assigned_agent) if ticket.assigned_agent else "",
    }


def _serialize_intervention(intervention):
    return {
        "ticket": intervention.ticket.reference,
        "technicien": str(intervention.agent),
        "client": str(intervention.ticket.client),
        "type": intervention.get_intervention_type_display(),
        "statut": intervention.get_status_display(),
        "debut": intervention.started_at.isoformat() if intervention.started_at else "",
        "fin": intervention.finished_at.isoformat() if intervention.finished_at else "",
        "action": intervention.action_taken,
    }


def _sla_compliance_rows(tickets):
    rows = []
    for priority, label in Ticket.PRIORITY_CHOICES:
        subset = tickets.filter(priority=priority)
        total = subset.count()
        if not total:
            rows.append({"priorite": label, "total": 0, "dans_les_delais": 0, "pourcentage": 0.0})
            continue
        compliant = 0
        for ticket in subset:
            if ticket.sla_deadline is None:
                continue
            if ticket.resolved_at:
                compliant += 1 if ticket.resolved_at <= ticket.sla_deadline else 0
            else:
                compliant += 1 if not ticket.is_overdue else 0
        rows.append(
            {
                "priorite": label,
                "total": total,
                "dans_les_delais": compliant,
                "pourcentage": round((compliant / total) * 100, 1),
            }
        )
    return rows


def _build_common_snapshot(user, report_type, anchor_date=None):
    start_date, end_date, start_dt, end_dt, label = _period_for(report_type, anchor_date=anchor_date)
    tickets = scope_ticket_queryset(
        Ticket.objects.select_related("client", "product", "assigned_agent").all(),
        user,
    )
    interventions = scope_intervention_queryset(
        Intervention.objects.select_related("ticket", "ticket__client", "agent").all(),
        user,
    )
    created_tickets = tickets.filter(created_at__gte=start_dt, created_at__lt=end_dt)
    resolved_tickets = tickets.filter(resolved_at__gte=start_dt, resolved_at__lt=end_dt)
    closed_tickets = tickets.filter(closed_at__gte=start_dt, closed_at__lt=end_dt)
    current_open = tickets.filter(status__in=OPEN_TICKET_STATUSES)
    overdue = current_open.filter(sla_deadline__lt=timezone.now())
    interventions_done = interventions.filter(
        Q(finished_at__gte=start_dt, finished_at__lt=end_dt) | Q(status=Intervention.STATUS_DONE, created_at__gte=start_dt, created_at__lt=end_dt)
    )

    return {
        "title": REPORT_CHOICES[report_type],
        "report_type": report_type,
        "period_label": label,
        "period_start": start_dt.isoformat(),
        "period_end": end_dt.isoformat(),
        "tickets_queryset": tickets,
        "created_queryset": created_tickets,
        "resolved_queryset": resolved_tickets,
        "closed_queryset": closed_tickets,
        "open_queryset": current_open,
        "overdue_queryset": overdue,
        "interventions_queryset": interventions_done,
        "summary": {
            "tickets_crees": created_tickets.count(),
            "tickets_resolus": resolved_tickets.count(),
            "tickets_ouverts": current_open.count(),
            "tickets_fermes": closed_tickets.count(),
            "tickets_hors_sla": overdue.count(),
            "premiere_reponse_moyenne_h": float(compute_average_first_response_hours(tickets) or 0),
            "resolution_moyenne_h": float(compute_average_resolution_hours(tickets) or 0),
        },
    }


def build_daily_report(user, anchor_date=None):
    snapshot = _build_common_snapshot(user, REPORT_DAILY, anchor_date=anchor_date)
    created_tickets = snapshot["created_queryset"]
    open_tickets = snapshot["open_queryset"]
    overdue = snapshot["overdue_queryset"]
    interventions_done = snapshot["interventions_queryset"]

    resolution_rate = 0.0
    if created_tickets.count():
        resolution_rate = round((snapshot["resolved_queryset"].count() / created_tickets.count()) * 100, 1)

    snapshot.update(
        {
            "tickets_par_statut": list(open_tickets.values("status").annotate(total=Count("id")).order_by("status")),
            "tickets_hors_sla_detail": [_serialize_ticket(ticket) for ticket in overdue[:20]],
            "interventions_realisees": [_serialize_intervention(item) for item in interventions_done[:30]],
            "resolution_rate_percent": resolution_rate,
            "tickets_non_assignes_30min": open_tickets.filter(
                assigned_agent__isnull=True,
                created_at__lte=timezone.now() - timedelta(minutes=30),
            ).count(),
            "tickets_nouveaux_1h": open_tickets.filter(
                status=Ticket.STATUS_NEW,
                created_at__lte=timezone.now() - timedelta(hours=1),
            ).count(),
        }
    )
    return snapshot


def build_weekly_report(user, anchor_date=None):
    snapshot = _build_common_snapshot(user, REPORT_WEEKLY, anchor_date=anchor_date)
    tickets = snapshot["tickets_queryset"]
    period_tickets = snapshot["created_queryset"]

    workload_curve = []
    current_day = datetime.fromisoformat(snapshot["period_start"]).date()
    end_day = datetime.fromisoformat(snapshot["period_end"]).date()
    while current_day < end_day:
        day_start, day_end = _aware_bounds(current_day, current_day + timedelta(days=1))
        workload_curve.append(
            {
                "jour": current_day.strftime("%a %d/%m"),
                "tickets_crees": period_tickets.filter(created_at__gte=day_start, created_at__lt=day_end).count(),
                "tickets_resolus": tickets.filter(resolved_at__gte=day_start, resolved_at__lt=day_end).count(),
            }
        )
        current_day += timedelta(days=1)

    snapshot.update(
        {
            "tickets_annules": period_tickets.filter(status=Ticket.STATUS_CANCELLED).count(),
            "performance_techniciens": compute_agent_performance_rows(tickets),
            "top_clients": list(
                period_tickets.values("client__username", "client__company_name")
                .annotate(total=Count("id"))
                .order_by("-total", "client__username")[:5]
            ),
            "top_equipements": list(
                period_tickets.values("product__name", "product__serial_number")
                .annotate(total=Count("id"))
                .exclude(product__isnull=True)
                .order_by("-total", "product__serial_number")[:5]
            ),
            "sla_par_priorite": _sla_compliance_rows(period_tickets),
            "courbe_charge": workload_curve,
        }
    )
    return snapshot


def build_monthly_report(user, anchor_date=None):
    snapshot = _build_common_snapshot(user, REPORT_MONTHLY, anchor_date=anchor_date)
    tickets = snapshot["tickets_queryset"]
    period_tickets = snapshot["created_queryset"]
    feedbacks = TicketFeedback.objects.filter(ticket__in=tickets).select_related("ticket")

    start_date = datetime.fromisoformat(snapshot["period_start"]).date()
    previous_anchor = start_date - timedelta(days=1)
    previous = _build_common_snapshot(user, REPORT_MONTHLY, anchor_date=previous_anchor)

    resolution_by_domain = []
    for domain, label in Ticket.BUSINESS_DOMAIN_CHOICES:
        domain_tickets = tickets.filter(business_domain=domain)
        average_hours = compute_average_resolution_hours(domain_tickets)
        resolution_by_domain.append(
            {
                "domaine": label,
                "tickets": domain_tickets.count(),
                "resolution_moyenne_h": float(average_hours) if average_hours is not None else 0.0,
            }
        )

    snapshot.update(
        {
            "temps_resolution_par_domaine": resolution_by_domain,
            "satisfaction_moyenne": float(feedbacks.aggregate(avg=Avg("rating"))["avg"] or 0),
            "comparaison_mois_precedent": {
                "tickets_crees": {
                    "courant": snapshot["summary"]["tickets_crees"],
                    "precedent": previous["summary"]["tickets_crees"],
                },
                "tickets_resolus": {
                    "courant": snapshot["summary"]["tickets_resolus"],
                    "precedent": previous["summary"]["tickets_resolus"],
                },
            },
            "causes_recurrentes": list(
                period_tickets.values("category", "business_domain").annotate(total=Count("id")).order_by("-total")[:10]
            ),
            "violations_sla_par_priorite": [
                {
                    "priorite": row["priorite"],
                    "violations": row["total"] - row["dans_les_delais"],
                    "total": row["total"],
                }
                for row in _sla_compliance_rows(period_tickets)
            ],
            "top_techniciens": compute_agent_performance_rows(tickets),
        }
    )
    return snapshot


REPORT_BUILDERS = {
    REPORT_DAILY: build_daily_report,
    REPORT_WEEKLY: build_weekly_report,
    REPORT_MONTHLY: build_monthly_report,
}


def build_report(report_type, user, anchor_date=None):
    try:
        builder = REPORT_BUILDERS[report_type]
    except KeyError as exc:
        raise ValueError("Type de rapport inconnu.") from exc
    report = builder(user, anchor_date=anchor_date)
    report.pop("tickets_queryset", None)
    report.pop("created_queryset", None)
    report.pop("resolved_queryset", None)
    report.pop("closed_queryset", None)
    report.pop("open_queryset", None)
    report.pop("overdue_queryset", None)
    report.pop("interventions_queryset", None)
    return report


def _render_section_rows(value):
    if isinstance(value, list) and value and isinstance(value[0], dict):
        headers = list(value[0].keys())
        rows = [headers]
        for item in value:
            rows.append([str(item.get(header, "")) for header in headers])
        return rows
    return None


def export_report_csv(report):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow([report.get("title", "Rapport"), report.get("period_label", "")])
    for key, value in report.items():
        if key in {"title", "report_type", "period_label", "period_start", "period_end"}:
            continue
        if isinstance(value, dict):
            writer.writerow([])
            writer.writerow([key])
            for sub_key, sub_value in value.items():
                writer.writerow([sub_key, sub_value])
        elif isinstance(value, list):
            rows = _render_section_rows(value)
            writer.writerow([])
            writer.writerow([key])
            if rows:
                writer.writerows(rows)
            else:
                for item in value:
                    writer.writerow([item])
        else:
            writer.writerow([key, value])
    return buffer.getvalue().encode("utf-8")


def export_report_xlsx(report):
    if Workbook is None:
        fallback_stream = StringIO()
        writer = csv.writer(fallback_stream, delimiter=";")
        writer.writerow(["Rapport", report.get("title", "")])
        writer.writerow(["Periode", report.get("period_label", "")])
        for key, value in report.items():
            if key in {"title", "report_type", "period_label", "period_start", "period_end"}:
                continue
            if isinstance(value, dict):
                writer.writerow([])
                writer.writerow([key])
                for sub_key, sub_value in value.items():
                    writer.writerow([sub_key, sub_value])
            elif isinstance(value, list):
                rows = _render_section_rows(value)
                writer.writerow([])
                writer.writerow([key])
                if rows:
                    writer.writerows(rows)
                else:
                    for item in value:
                        writer.writerow([item])
            else:
                writer.writerow([key, value])
        return fallback_stream.getvalue().encode("utf-8")

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resume"
    summary_sheet.append(["Rapport", report.get("title", "")])
    summary_sheet.append(["Periode", report.get("period_label", "")])

    row_index = 4
    for key, value in report.items():
        if key in {"title", "report_type", "period_label", "period_start", "period_end"}:
            continue
        if isinstance(value, dict):
            summary_sheet.cell(row=row_index, column=1, value=key)
            row_index += 1
            for sub_key, sub_value in value.items():
                summary_sheet.append([sub_key, sub_value])
                row_index += 1
        elif isinstance(value, list):
            rows = _render_section_rows(value)
            if rows:
                sheet = workbook.create_sheet(title=key[:31])
                for row in rows:
                    sheet.append(row)
            else:
                summary_sheet.append([key, ", ".join(str(item) for item in value)])
                row_index += 1
        else:
            summary_sheet.append([key, value])
            row_index += 1

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def export_report_pdf(report):
    stream = BytesIO()
    document = SimpleDocTemplate(stream, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(report.get("title", "Rapport"), styles["Title"]),
        Paragraph(f"Periode: {report.get('period_label', '')}", styles["Normal"]),
        Spacer(1, 12),
    ]

    for key, value in report.items():
        if key in {"title", "report_type", "period_label", "period_start", "period_end"}:
            continue
        story.append(Paragraph(str(key).replace("_", " ").title(), styles["Heading3"]))
        if isinstance(value, dict):
            for sub_key, sub_value in value.items():
                story.append(Paragraph(f"{sub_key}: {sub_value}", styles["Normal"]))
        elif isinstance(value, list):
            rows = _render_section_rows(value)
            if rows:
                table = Table(rows, repeatRows=1)
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8D5BF")),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ]
                    )
                )
                story.append(table)
            else:
                for item in value:
                    story.append(Paragraph(str(item), styles["Normal"]))
        else:
            story.append(Paragraph(str(value), styles["Normal"]))
        story.append(Spacer(1, 10))

    document.build(story)
    return stream.getvalue()
